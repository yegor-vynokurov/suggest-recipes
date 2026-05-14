from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.kitchen.inventory_excel import (
    DEFAULT_INGREDIENTS_DIR,
    DOMAIN_TITLES,
    make_inventory_excel,
)


def build_sheet_rows_by_item_id(workbook_path: Path, sheet_name: str) -> dict[str, dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    worksheet = workbook[sheet_name]

    headers = [cell.value for cell in worksheet[1]]
    rows: dict[str, dict[str, str]] = {}

    for values in worksheet.iter_rows(min_row=2, values_only=True):
        if not values or values[1] is None:
            continue

        row = {
            header: ("" if value is None else str(value))
            for header, value in zip(headers, values, strict=False)
        }
        rows[row["item_id"]] = row

    return rows


def test_make_inventory_excel_includes_recently_added_ingredients(tmp_path: Path) -> None:
    inventory_xlsx = tmp_path / "inventory.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    existing_csv.write_text(
        "\n".join(
            [
                "amount,item_id",
                "many,langoustine",
                "normal,salami",
                "little,lefse",
                "normal,apricot_jam",
                "many,raspberry",
                "normal,mozzarella",
                "normal,chestnuts",
                "little,coffee",
            ]
        ),
        encoding="utf-8",
    )

    result_path = make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
    )

    assert result_path == inventory_xlsx
    assert inventory_xlsx.exists()

    workbook = load_workbook(inventory_xlsx, data_only=True)
    assert "README" in workbook.sheetnames
    assert "ingredients_uncategorized" not in workbook.sheetnames

    fish_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_fish_seafood"])
    meat_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_meat_poultry"])
    dairy_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_dairy_eggs"])
    bread_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_grains_bread"])
    fruit_rows = build_sheet_rows_by_item_id(
        inventory_xlsx,
        DOMAIN_TITLES["ingredients_vegetables_mushrooms_fruits"],
    )
    sweet_rows = build_sheet_rows_by_item_id(
        inventory_xlsx,
        DOMAIN_TITLES["ingredients_nuts_seeds_sweeteners"],
    )
    sauce_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_sauces_condiments"])
    spice_rows = build_sheet_rows_by_item_id(inventory_xlsx, DOMAIN_TITLES["ingredients_spices_herbs"])

    assert fish_rows["langoustine"]["amount"] == "many"
    assert "crustacean" in fish_rows["langoustine"]["groups"]
    assert fish_rows["langoustine"]["source_file"] == "ingredients_fish_seafood.yaml"

    assert meat_rows["salami"]["amount"] == "normal"
    assert "prepared_meat" in meat_rows["salami"]["groups"]
    assert meat_rows["salami"]["source_file"] == "ingredients_meat_poultry.yaml"

    assert dairy_rows["mozzarella"]["amount"] == "normal"
    assert "dairy" in dairy_rows["mozzarella"]["groups"]
    assert dairy_rows["mozzarella"]["source_file"] == "ingredients_dairy_eggs.yaml"

    assert bread_rows["lefse"]["amount"] == "little"
    assert "flatbread" in bread_rows["lefse"]["groups"]
    assert bread_rows["lefse"]["source_file"] == "ingredients_grains_bread.yaml"

    assert sauce_rows["apricot_jam"]["amount"] == "normal"
    assert "fruit_preserve" in sauce_rows["apricot_jam"]["groups"]
    assert sauce_rows["apricot_jam"]["source_file"] == "ingredients_sauces_condiments.yaml"

    assert fruit_rows["raspberry"]["amount"] == "many"
    assert "berry" in fruit_rows["raspberry"]["groups"]
    assert fruit_rows["raspberry"]["source_file"] == "ingredients_vegetables_mushrooms_fruits.yaml"

    assert sweet_rows["chestnuts"]["amount"] == "normal"
    assert "nuts" in sweet_rows["chestnuts"]["groups"]
    assert sweet_rows["chestnuts"]["source_file"] == "ingredients_nuts_seeds_sweeteners.yaml"

    assert spice_rows["coffee"]["amount"] == "little"
    assert "warm_spice" in spice_rows["coffee"]["groups"]
    assert spice_rows["coffee"]["source_file"] == "ingredients_spices_herbs.yaml"
