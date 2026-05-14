from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.kitchen.i18n import t
from src.kitchen.inventory_excel import (
    DEFAULT_INGREDIENTS_DIR,
    build_parser,
    default_inventory_csv_for_lang,
    default_inventory_xlsx_for_lang,
    domain_title,
    export_all_default_inventory_excels_to_csv,
    export_excel_to_current_inventory_csv,
    localized_headers,
    make_inventory_excel,
    readme_sheet_title,
)


def test_default_inventory_xlsx_for_lang_uses_language_specific_names() -> None:
    assert default_inventory_xlsx_for_lang("en") == Path("data/inventory/current_inventory.en.xlsx")
    assert default_inventory_xlsx_for_lang("uk") == Path("data/inventory/current_inventory.uk.xlsx")


def test_default_inventory_csv_for_lang_uses_language_specific_names() -> None:
    assert default_inventory_csv_for_lang("en") == Path("data/inventory/current_inventory.en.csv")
    assert default_inventory_csv_for_lang("uk") == Path("data/inventory/current_inventory.uk.csv")


def test_inventory_excel_parser_accepts_lang_before_subcommand() -> None:
    parser = build_parser()

    args = parser.parse_args(["--lang", "uk", "make-excel"])

    assert args.lang == "uk"
    assert args.inventory_xlsx is None


def test_inventory_excel_parser_accepts_lang_after_subcommand() -> None:
    parser = build_parser()

    args = parser.parse_args(["export-csv", "--lang", "uk"])

    assert args.lang == "uk"
    assert args.inventory_xlsx is None


def test_make_inventory_excel_uses_language_specific_default_path(
    monkeypatch,
    tmp_path: Path,
) -> None:
    repo_root = Path.cwd()
    ingredients_dir = repo_root / DEFAULT_INGREDIENTS_DIR
    existing_csv = tmp_path / "current_inventory.csv"
    existing_csv.write_text("amount,item_id\n", encoding="utf-8")

    monkeypatch.chdir(tmp_path)

    result_path = make_inventory_excel(
        ingredients_dir=ingredients_dir,
        inventory_xlsx=None,
        existing_csv=existing_csv,
        lang="uk",
    )

    assert result_path == Path("data/inventory/current_inventory.uk.xlsx")
    assert (tmp_path / result_path).exists()


def test_inventory_excel_uses_localized_aliases_for_selected_language(tmp_path: Path) -> None:
    inventory_xlsx = tmp_path / "inventory.uk.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    existing_csv.write_text("amount,item_id\nnormal,mustard\n", encoding="utf-8")

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang="uk",
    )

    workbook = load_workbook(inventory_xlsx, data_only=True)
    sheet = workbook[domain_title("ingredients_sauces_condiments", "uk")]
    header_index = {cell.value: idx for idx, cell in enumerate(sheet[1], start=1)}

    row_index = None
    for candidate_row in range(2, sheet.max_row + 1):
        if sheet.cell(row=candidate_row, column=header_index["item_id"]).value == "mustard":
            row_index = candidate_row
            break

    assert row_index is not None
    assert sheet.cell(row=row_index, column=header_index[t("excel.header.aliases", "uk")]).value == "гірчиця"


@pytest.mark.parametrize("lang", ["en", "uk"])
def test_make_inventory_excel_localizes_sheet_titles_readme_headers_and_validation(
    tmp_path: Path,
    lang: str,
) -> None:
    inventory_xlsx = tmp_path / f"inventory.{lang}.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    existing_csv.write_text("amount,item_id\nnormal,mozzarella\n", encoding="utf-8")

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang=lang,
    )

    workbook = load_workbook(inventory_xlsx, data_only=True)

    assert readme_sheet_title(lang) in workbook.sheetnames
    assert domain_title("ingredients_dairy_eggs", lang) in workbook.sheetnames

    readme = workbook[readme_sheet_title(lang)]
    assert readme["A1"].value == t("excel.readme.title", lang)
    assert readme["A3"].value == t("excel.readme.line_1", lang)
    assert readme["A4"].value == t("excel.readme.line_2", lang)
    assert readme["A5"].value == t("excel.readme.line_3", lang)
    assert readme["A6"].value == t("excel.readme.line_4", lang)
    assert readme["A8"].value == t("excel.readme.amount_values_title", lang)

    dairy_sheet = workbook[domain_title("ingredients_dairy_eggs", lang)]
    header_values = [cell.value for cell in dairy_sheet[1]]
    assert header_values == localized_headers(lang)

    validation = next(iter(dairy_sheet.data_validations.dataValidation))
    assert validation.error == t("excel.validation.amount_error", lang)
    assert validation.errorTitle == t("excel.validation.amount_error_title", lang)
    assert validation.prompt == t("excel.validation.amount_prompt", lang)
    assert validation.promptTitle == t("excel.validation.amount_prompt_title", lang)


@pytest.mark.parametrize("lang", ["en", "uk"])
def test_inventory_excel_roundtrip_uses_item_id_across_languages(
    tmp_path: Path,
    lang: str,
) -> None:
    inventory_xlsx = tmp_path / f"inventory.{lang}.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    inventory_csv = tmp_path / "exported.csv"
    existing_csv.write_text(
        "amount,item_id\nmany,langoustine\nnormal,mozzarella\n",
        encoding="utf-8",
    )

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang=lang,
    )

    workbook = load_workbook(inventory_xlsx)
    dairy_sheet = workbook[domain_title("ingredients_dairy_eggs", lang)]
    seafood_sheet = workbook[domain_title("ingredients_fish_seafood", lang)]

    dairy_header_index = {cell.value: idx for idx, cell in enumerate(dairy_sheet[1], start=1)}
    seafood_header_index = {cell.value: idx for idx, cell in enumerate(seafood_sheet[1], start=1)}

    dairy_name_col = dairy_header_index[localized_headers(lang)[2]]
    seafood_name_col = seafood_header_index[localized_headers(lang)[2]]

    dairy_item_id_col = dairy_header_index["item_id"]
    seafood_item_id_col = seafood_header_index["item_id"]

    dairy_row = None
    for row_index in range(2, dairy_sheet.max_row + 1):
        if dairy_sheet.cell(row=row_index, column=dairy_item_id_col).value == "mozzarella":
            dairy_row = row_index
            break

    seafood_row = None
    for row_index in range(2, seafood_sheet.max_row + 1):
        if seafood_sheet.cell(row=row_index, column=seafood_item_id_col).value == "langoustine":
            seafood_row = row_index
            break

    assert dairy_row is not None
    assert seafood_row is not None

    dairy_sheet.cell(row=dairy_row, column=dairy_name_col, value=f"Wrong {lang} dairy name")
    seafood_sheet.cell(row=seafood_row, column=seafood_name_col, value=f"Wrong {lang} seafood name")
    workbook.save(inventory_xlsx)

    export_excel_to_current_inventory_csv(
        inventory_xlsx=inventory_xlsx,
        inventory_csv=inventory_csv,
        lang=lang,
    )

    csv_lines = inventory_csv.read_text(encoding="utf-8").splitlines()
    assert csv_lines[0] == "item_id,amount"
    assert "name" not in csv_lines[0]

    amounts_by_item_id = {}
    for line in csv_lines[1:]:
        item_id, amount = line.split(",", maxsplit=1)
        amounts_by_item_id[item_id] = amount

    assert amounts_by_item_id["langoustine"] == "many"
    assert amounts_by_item_id["mozzarella"] == "normal"
    assert not any("Wrong " in line for line in csv_lines)


def test_export_excel_to_current_inventory_csv_ignores_edited_name_column(
    tmp_path: Path,
) -> None:
    inventory_xlsx = tmp_path / "inventory.en.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    inventory_csv = tmp_path / "exported.csv"
    existing_csv.write_text("amount,item_id\nnormal,mozzarella\n", encoding="utf-8")

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang="en",
    )

    workbook = load_workbook(inventory_xlsx)
    dairy_sheet = workbook[domain_title("ingredients_dairy_eggs", "en")]
    header_index = {cell.value: idx for idx, cell in enumerate(dairy_sheet[1], start=1)}
    dairy_sheet.cell(row=2, column=header_index["name"], value="Totally wrong name")
    workbook.save(inventory_xlsx)

    export_excel_to_current_inventory_csv(
        inventory_xlsx=inventory_xlsx,
        inventory_csv=inventory_csv,
        lang="en",
    )

    csv_text = inventory_csv.read_text(encoding="utf-8")
    assert "Totally wrong name" not in csv_text
    assert "item_id,amount" in csv_text
    assert "mozzarella,normal" in csv_text


def test_export_excel_to_current_inventory_csv_requires_item_id_value(
    tmp_path: Path,
) -> None:
    inventory_xlsx = tmp_path / "inventory.en.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    inventory_csv = tmp_path / "exported.csv"
    existing_csv.write_text("amount,item_id\nnormal,mozzarella\n", encoding="utf-8")

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang="en",
    )

    workbook = load_workbook(inventory_xlsx)
    dairy_sheet = workbook[domain_title("ingredients_dairy_eggs", "en")]
    header_index = {cell.value: idx for idx, cell in enumerate(dairy_sheet[1], start=1)}
    target_row = None
    for row_index in range(2, dairy_sheet.max_row + 1):
        if dairy_sheet.cell(row=row_index, column=header_index["item_id"]).value == "mozzarella":
            target_row = row_index
            break

    assert target_row is not None
    dairy_sheet.cell(row=target_row, column=header_index["item_id"], value="")
    dairy_sheet.cell(row=target_row, column=header_index["amount"], value="many")
    workbook.save(inventory_xlsx)

    export_excel_to_current_inventory_csv(
        inventory_xlsx=inventory_xlsx,
        inventory_csv=inventory_csv,
        lang="en",
    )

    csv_lines = inventory_csv.read_text(encoding="utf-8").splitlines()
    assert csv_lines[0] == "item_id,amount"
    assert not any(line.startswith("mozzarella,") for line in csv_lines[1:])


def test_export_all_default_inventory_excels_to_csv_exports_all_existing_language_files(
    monkeypatch,
    tmp_path: Path,
) -> None:
    repo_root = Path.cwd()
    ingredients_dir = repo_root / DEFAULT_INGREDIENTS_DIR
    monkeypatch.chdir(tmp_path)

    for lang, item_id in (("en", "mozzarella"), ("uk", "langoustine")):
        existing_csv = tmp_path / f"seed.{lang}.csv"
        existing_csv.write_text(f"amount,item_id\nnormal,{item_id}\n", encoding="utf-8")
        make_inventory_excel(
            ingredients_dir=ingredients_dir,
            inventory_xlsx=default_inventory_xlsx_for_lang(lang),
            existing_csv=existing_csv,
            lang=lang,
        )

    exported = export_all_default_inventory_excels_to_csv()

    assert exported == {
        "en": Path("data/inventory/current_inventory.en.csv"),
        "uk": Path("data/inventory/current_inventory.uk.csv"),
    }
    assert (tmp_path / exported["en"]).exists()
    assert (tmp_path / exported["uk"]).exists()

    en_lines = (tmp_path / exported["en"]).read_text(encoding="utf-8").splitlines()
    uk_lines = (tmp_path / exported["uk"]).read_text(encoding="utf-8").splitlines()

    assert en_lines[0] == "item_id,amount"
    assert uk_lines[0] == "item_id,amount"
    assert any(line.startswith("mozzarella,normal") for line in en_lines[1:])
    assert any(line.startswith("langoustine,normal") for line in uk_lines[1:])


def test_make_inventory_excel_prints_localized_summary_in_english(
    tmp_path: Path,
    capsys,
) -> None:
    inventory_xlsx = tmp_path / "inventory.en.xlsx"
    existing_csv = tmp_path / "current_inventory.csv"
    existing_csv.write_text("amount,item_id\nnormal,mozzarella\n", encoding="utf-8")

    make_inventory_excel(
        ingredients_dir=DEFAULT_INGREDIENTS_DIR,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
        lang="en",
    )

    output = capsys.readouterr().out
    assert "Inventory Excel template ready" in output
    assert "Product sheets" in output
    assert "Ingredients to fill in manually" in output
    assert "кладовки" not in output
    assert "Строк записано" not in output


def test_export_all_default_inventory_excels_to_csv_uses_requested_output_language(
    monkeypatch,
    tmp_path: Path,
    capsys,
) -> None:
    repo_root = Path.cwd()
    ingredients_dir = repo_root / DEFAULT_INGREDIENTS_DIR
    monkeypatch.chdir(tmp_path)

    for lang, item_id in (("en", "mozzarella"), ("uk", "langoustine")):
        existing_csv = tmp_path / f"seed.{lang}.csv"
        existing_csv.write_text(f"amount,item_id\nnormal,{item_id}\n", encoding="utf-8")
        make_inventory_excel(
            ingredients_dir=ingredients_dir,
            inventory_xlsx=default_inventory_xlsx_for_lang(lang),
            existing_csv=existing_csv,
            lang=lang,
            verbose=False,
        )

    export_all_default_inventory_excels_to_csv(output_lang="en")
    output = capsys.readouterr().out

    assert output.count("Inventory CSV ready") == 2
    assert "Rows written" in output
    assert "CSV кладовки готов" not in output
