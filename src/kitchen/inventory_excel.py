from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from .i18n import DEFAULT_LANG, SUPPORTED_LANGUAGES, normalize_language, t
from .yaml_store import DEFAULT_YAML_STORE, YamlStore


AMOUNT_VALUES = ["none", "spice", "little", "normal", "many"]

DEFAULT_INGREDIENTS_DIR = Path("data/inventory/ingredients")
DEFAULT_INVENTORY_XLSX_EN = Path("data/inventory/current_inventory.en.xlsx")
DEFAULT_INVENTORY_XLSX_UK = Path("data/inventory/current_inventory.uk.xlsx")
DEFAULT_INVENTORY_CSV_EN = Path("data/inventory/current_inventory.en.csv")
DEFAULT_INVENTORY_CSV_UK = Path("data/inventory/current_inventory.uk.csv")
DEFAULT_INVENTORY_CSV_LEGACY = Path("data/inventory/current_inventory.csv")
DEFAULT_INVENTORY_XLSX_BY_LANG = {
    "en": DEFAULT_INVENTORY_XLSX_EN,
    "uk": DEFAULT_INVENTORY_XLSX_UK,
}
DEFAULT_INVENTORY_CSV_BY_LANG = {
    "en": DEFAULT_INVENTORY_CSV_EN,
    "uk": DEFAULT_INVENTORY_CSV_UK,
}

DOMAIN_KEYS = [
    "ingredients_meat_poultry",
    "ingredients_fish_seafood",
    "ingredients_dairy_eggs",
    "ingredients_legumes",
    "ingredients_grains_bread",
    "ingredients_vegetables_mushrooms_fruits",
    "ingredients_fats_oils",
    "ingredients_nuts_seeds_sweeteners",
    "ingredients_sauces_condiments",
    "ingredients_spices_herbs",
    "ingredients_technical",
    "ingredients_prepared_components",
]
DOMAIN_TITLES = {domain_key: t(f"excel.domain.{domain_key}", "en") for domain_key in DOMAIN_KEYS}

SKIP_FILENAMES = {
    "ingredients_all_merged.yaml",
    "ingredient_categories.yaml",
    "ingredients_abstract.yaml",
}

HEADERS = [
    "amount",
    "item_id",
    "name",
    "groups",
    "aliases",
    "source_file",
    "review",
    "comment",
]


def domain_title(domain_key: str, lang: str | None = DEFAULT_LANG) -> str:
    translation_key = f"excel.domain.{domain_key}"
    translated = t(translation_key, lang)
    return domain_key if translated == translation_key else translated


def readme_sheet_title(lang: str | None = DEFAULT_LANG) -> str:
    translation_key = "excel.sheet.readme"
    translated = t(translation_key, lang)
    return "README" if translated == translation_key else translated


def localized_headers(lang: str | None = DEFAULT_LANG) -> list[str]:
    labels = []
    for header_key in HEADERS:
        translation_key = f"excel.header.{header_key}"
        translated = t(translation_key, lang)
        labels.append(header_key if translated == translation_key else translated)
    return labels


def header_aliases() -> dict[str, str]:
    aliases = {header_key: header_key for header_key in HEADERS}

    for lang in SUPPORTED_LANGUAGES:
        for header_key in HEADERS:
            aliases[t(f"excel.header.{header_key}", lang)] = header_key

    return aliases


def localized_item_text(
    item: dict[str, Any],
    field: str,
    lang: str | None = DEFAULT_LANG,
    fallback: str = "",
) -> str:
    normalized_lang = normalize_language(lang)
    value = item.get(field)

    if isinstance(value, dict):
        localized_value = value.get(normalized_lang) or value.get(DEFAULT_LANG)
        if isinstance(localized_value, str) and localized_value.strip():
            return localized_value.strip()

    if normalized_lang == "uk":
        uk_value = item.get(f"{field}_uk")
        if isinstance(uk_value, str) and uk_value.strip():
            return uk_value.strip()

    if isinstance(value, str) and value.strip():
        return value.strip()

    return fallback


def localized_aliases(
    item: dict[str, Any],
    lang: str | None = DEFAULT_LANG,
) -> list[str]:
    normalized_lang = normalize_language(lang)

    if normalized_lang == "uk":
        uk_aliases = item.get("aliases_uk")
        if isinstance(uk_aliases, list):
            return [
                alias.strip()
                for alias in uk_aliases
                if isinstance(alias, str) and alias.strip()
            ]

    aliases = item.get("aliases")
    if isinstance(aliases, list):
        return [
            alias.strip()
            for alias in aliases
            if isinstance(alias, str) and alias.strip()
        ]

    return []


def default_inventory_xlsx_for_lang(lang: str | None = DEFAULT_LANG) -> Path:
    return DEFAULT_INVENTORY_XLSX_BY_LANG[normalize_language(lang)]


def default_inventory_csv_for_lang(lang: str | None = DEFAULT_LANG) -> Path:
    return DEFAULT_INVENTORY_CSV_BY_LANG[normalize_language(lang)]


def is_inventory_visible(item_id: str, item: dict[str, Any]) -> bool:
    inventory = item.get("inventory", {}) or {}

    if item.get("abstract"):
        return False

    if inventory.get("track") is False:
        return False

    if inventory.get("hide_from_template") is True:
        return False

    return True


def default_amount_for(item: dict[str, Any]) -> str:
    inventory = item.get("inventory", {}) or {}
    amount = inventory.get("default_amount", "none")

    if amount not in AMOUNT_VALUES:
        return "none"

    return amount


def safe_sheet_name(raw_name: str, used: set[str], lang: str | None = DEFAULT_LANG) -> str:
    name = domain_title(raw_name, lang)
    name = re.sub(r"[\[\]\:\*\?\/\\]", "_", name)
    name = name[:31] or "Sheet"

    base = name
    index = 2
    while name in used:
        suffix = f"_{index}"
        name = f"{base[:31-len(suffix)]}{suffix}"
        index += 1

    used.add(name)
    return name


@dataclass
class InventoryExcelService:
    lang: str = DEFAULT_LANG
    ingredients_dir: Path = DEFAULT_INGREDIENTS_DIR
    inventory_xlsx: Path | None = None
    existing_csv: Path | None = None
    inventory_csv: Path | None = None
    yaml_store: YamlStore = field(default_factory=lambda: DEFAULT_YAML_STORE)

    def __post_init__(self) -> None:
        self.lang = normalize_language(self.lang)
        self.ingredients_dir = Path(self.ingredients_dir)
        self.inventory_xlsx = (
            Path(self.inventory_xlsx)
            if self.inventory_xlsx is not None
            else default_inventory_xlsx_for_lang(self.lang)
        )
        self.existing_csv = (
            Path(self.existing_csv)
            if self.existing_csv is not None
            else default_inventory_csv_for_lang(self.lang)
        )
        self.inventory_csv = (
            Path(self.inventory_csv)
            if self.inventory_csv is not None
            else default_inventory_csv_for_lang(self.lang)
        )

    def load_yaml(self, path: Path) -> dict[str, Any]:
        return self.yaml_store.load(path)

    def read_existing_csv_amounts(self, path: Path | None = None) -> dict[str, str]:
        candidate_paths: list[Path]
        if path is not None:
            candidate_paths = [Path(path)]
        else:
            candidate_paths = [self.existing_csv]
            if (
                self.existing_csv == default_inventory_csv_for_lang(self.lang)
                and DEFAULT_INVENTORY_CSV_LEGACY not in candidate_paths
            ):
                candidate_paths.append(DEFAULT_INVENTORY_CSV_LEGACY)

        csv_path = next((candidate for candidate in candidate_paths if candidate.exists()), None)
        if csv_path is None:
            return {}

        text = csv_path.read_text(encoding="utf-8-sig")
        if not text.strip():
            return {}

        sample = text[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel

        amounts: dict[str, str] = {}

        with csv_path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file, dialect=dialect)
            for row in reader:
                item_id = (row.get("item_id") or row.get("id") or "").strip()
                amount = (row.get("amount") or "").strip()

                if item_id and amount in AMOUNT_VALUES:
                    amounts[item_id] = amount

        return amounts

    def iter_ingredient_files(self) -> list[Path]:
        files = []

        for path in sorted(self.ingredients_dir.glob("ingredients_*.yaml")):
            if path.name in SKIP_FILENAMES:
                continue
            files.append(path)

        return files

    def rows_from_ingredient_file(
        self,
        path: Path,
        existing_amounts: dict[str, str],
    ) -> list[dict[str, str]]:
        data = self.load_yaml(path)
        ingredients = data.get("ingredients", {}) or {}

        rows = []
        for item_id, item in sorted(ingredients.items(), key=lambda pair: pair[1].get("name", pair[0])):
            if not is_inventory_visible(item_id, item):
                continue

            amount = existing_amounts.get(item_id, default_amount_for(item))
            rows.append({
                "amount": amount,
                "item_id": item_id,
                "name": localized_item_text(item, "name", self.lang, fallback=item_id),
                "groups": ", ".join(item.get("groups", []) or []),
                "aliases": ", ".join(localized_aliases(item, self.lang)),
                "source_file": path.name,
                "review": "yes" if item.get("review") else "",
                "comment": localized_item_text(item, "comment", self.lang),
            })

        return rows

    def style_inventory_sheet(self, ws, row_count: int) -> None:
        header_fill = PatternFill("solid", fgColor="386641")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2D0")
        border = Border(bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        widths = {
            "A": 13,
            "B": 26,
            "C": 34,
            "D": 42,
            "E": 28,
            "F": 30,
            "G": 10,
            "H": 42,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{max(row_count, 1)}"

        for row in ws.iter_rows(min_row=2, max_row=max(row_count, 2), min_col=1, max_col=8):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        if row_count >= 2:
            validation = DataValidation(
                type="list",
                formula1='"' + ",".join(AMOUNT_VALUES) + '"',
                allow_blank=False,
            )
            validation.error = t("excel.validation.amount_error", self.lang)
            validation.errorTitle = t("excel.validation.amount_error_title", self.lang)
            validation.prompt = t("excel.validation.amount_prompt", self.lang)
            validation.promptTitle = t("excel.validation.amount_prompt_title", self.lang)
            ws.add_data_validation(validation)
            validation.add(f"A2:A{row_count}")

        amount_fills = {
            "none": "F2F2F2",
            "spice": "E8F4FD",
            "little": "FFF4CC",
            "normal": "E6F4EA",
            "many": "D9EAD3",
        }
        for row_index in range(2, row_count + 1):
            amount = ws.cell(row=row_index, column=1).value
            fill = amount_fills.get(amount)
            if fill:
                ws.cell(row=row_index, column=1).fill = PatternFill("solid", fgColor=fill)

    def build_workbook(self) -> tuple[Workbook, int, int]:
        if not self.ingredients_dir.exists():
            raise FileNotFoundError(
                t("excel.error.ingredients_dir_not_found", self.lang, path=self.ingredients_dir)
            )

        existing_amounts = self.read_existing_csv_amounts()
        wb = Workbook()
        wb.remove(wb.active)

        readme = wb.create_sheet(readme_sheet_title(self.lang))
        readme["A1"] = t("excel.readme.title", self.lang)
        readme["A1"].font = Font(bold=True, size=14)
        readme["A3"] = t("excel.readme.line_1", self.lang)
        readme["A4"] = t("excel.readme.line_2", self.lang)
        readme["A5"] = t("excel.readme.line_3", self.lang)
        readme["A6"] = t("excel.readme.line_4", self.lang)
        readme["A8"] = t("excel.readme.amount_values_title", self.lang)
        for idx, value in enumerate(AMOUNT_VALUES, start=9):
            readme.cell(row=idx, column=1, value=value)
        readme.column_dimensions["A"].width = 90

        used_sheet_names = {readme_sheet_title(self.lang)}
        total_rows = 0
        created_sheets = 0

        for path in self.iter_ingredient_files():
            rows = self.rows_from_ingredient_file(path, existing_amounts)
            if not rows:
                continue

            ws = wb.create_sheet(safe_sheet_name(path.stem, used_sheet_names, self.lang))
            ws.append(localized_headers(self.lang))
            for row in rows:
                ws.append([row[header] for header in HEADERS])

            row_count = len(rows) + 1
            self.style_inventory_sheet(ws, row_count)

            total_rows += len(rows)
            created_sheets += 1

        if created_sheets == 0:
            raise ValueError(
                t("excel.error.no_visible_ingredients", self.lang, path=self.ingredients_dir)
            )

        return wb, created_sheets, total_rows

    def make_inventory_excel(
        self,
        inventory_xlsx: Path | None = None,
        *,
        verbose: bool = True,
    ) -> Path:
        if inventory_xlsx is not None:
            self.inventory_xlsx = Path(inventory_xlsx)

        wb, created_sheets, total_rows = self.build_workbook()
        self.inventory_xlsx.parent.mkdir(parents=True, exist_ok=True)
        wb.save(self.inventory_xlsx)

        if verbose:
            print(f"{t('excel.message.inventory_excel_ready', self.lang)}: {self.inventory_xlsx}")
            print(f"{t('excel.message.product_sheet_count', self.lang)}: {created_sheets}")
            print(f"{t('excel.message.manual_fill_ingredient_count', self.lang)}: {total_rows}")

        return self.inventory_xlsx

    def read_inventory_excel_rows(self, inventory_xlsx: Path | None = None) -> list[dict[str, str]]:
        workbook_path = Path(inventory_xlsx) if inventory_xlsx is not None else self.inventory_xlsx
        wb = load_workbook(workbook_path, data_only=True)

        rows: list[dict[str, str]] = []
        localized_readme_titles = {readme_sheet_title(lang) for lang in SUPPORTED_LANGUAGES}
        localized_header_aliases = header_aliases()

        for ws in wb.worksheets:
            if ws.title.startswith("_") or ws.title in localized_readme_titles:
                continue

            headers = [
                localized_header_aliases.get(str(cell.value).strip(), str(cell.value).strip())
                if cell.value is not None
                else None
                for cell in ws[1]
            ]
            if not headers or "item_id" not in headers or "amount" not in headers:
                continue

            index = {name: pos for pos, name in enumerate(headers) if name}

            for values in ws.iter_rows(min_row=2, values_only=True):
                item_id = values[index["item_id"]]
                amount = values[index["amount"]]

                if item_id is None:
                    continue

                item_id = str(item_id).strip()
                amount = str(amount or "none").strip()

                if not item_id:
                    continue

                if amount not in AMOUNT_VALUES:
                    amount = "none"

                rows.append({
                    "item_id": item_id,
                    "amount": amount,
                })

        merged: dict[str, dict[str, str]] = {}
        for row in rows:
            item_id = row["item_id"]
            if item_id not in merged:
                merged[item_id] = row
            elif merged[item_id]["amount"] == "none" and row["amount"] != "none":
                merged[item_id] = row

        return sorted(merged.values(), key=lambda row: row["item_id"])

    def export_excel_to_current_inventory_csv(
        self,
        inventory_xlsx: Path | None = None,
        inventory_csv: Path | None = None,
        include_none: bool = True,
        *,
        verbose: bool = True,
    ) -> Path:
        if inventory_xlsx is not None:
            self.inventory_xlsx = Path(inventory_xlsx)
        if inventory_csv is not None:
            self.inventory_csv = Path(inventory_csv)

        if not self.inventory_xlsx.exists():
            raise FileNotFoundError(
                t("excel.error.inventory_excel_not_found", self.lang, path=self.inventory_xlsx)
            )

        rows = self.read_inventory_excel_rows(self.inventory_xlsx)
        if not include_none:
            rows = [row for row in rows if row["amount"] != "none"]

        self.inventory_csv.parent.mkdir(parents=True, exist_ok=True)
        with self.inventory_csv.open("w", encoding="utf-8", newline="") as file:
            writer = csv.DictWriter(file, fieldnames=["item_id", "amount"])
            writer.writeheader()
            writer.writerows(rows)

        if verbose:
            print(f"{t('excel.message.inventory_csv_ready', self.lang)}: {self.inventory_csv}")
            print(f"{t('excel.message.rows_written', self.lang)}: {len(rows)}")
        return self.inventory_csv


def make_inventory_excel(
    ingredients_dir: Path = DEFAULT_INGREDIENTS_DIR,
    inventory_xlsx: Path | None = None,
    existing_csv: Path | None = None,
    lang: str = DEFAULT_LANG,
    *,
    verbose: bool = True,
) -> Path:
    service = InventoryExcelService(
        lang=lang,
        ingredients_dir=ingredients_dir,
        inventory_xlsx=inventory_xlsx,
        existing_csv=existing_csv,
    )
    return service.make_inventory_excel(inventory_xlsx=inventory_xlsx, verbose=verbose)


def read_inventory_excel_rows(inventory_xlsx: Path) -> list[dict[str, str]]:
    return InventoryExcelService(inventory_xlsx=inventory_xlsx).read_inventory_excel_rows(inventory_xlsx)


def export_excel_to_current_inventory_csv(
    inventory_xlsx: Path | None = None,
    inventory_csv: Path | None = None,
    include_none: bool = True,
    lang: str = DEFAULT_LANG,
    *,
    verbose: bool = True,
) -> Path:
    service = InventoryExcelService(
        lang=lang,
        inventory_xlsx=inventory_xlsx,
        inventory_csv=inventory_csv,
    )
    return service.export_excel_to_current_inventory_csv(
        inventory_xlsx=inventory_xlsx,
        inventory_csv=inventory_csv,
        include_none=include_none,
        verbose=verbose,
    )


def export_all_default_inventory_excels_to_csv(
    include_none: bool = True,
    output_lang: str = DEFAULT_LANG,
    *,
    verbose: bool = True,
) -> dict[str, Path]:
    exported: dict[str, Path] = {}
    checked_paths: list[Path] = []

    for lang in SUPPORTED_LANGUAGES:
        inventory_xlsx = default_inventory_xlsx_for_lang(lang)
        inventory_csv = default_inventory_csv_for_lang(lang)
        checked_paths.append(inventory_xlsx)

        if not inventory_xlsx.exists():
            continue

        exported[lang] = export_excel_to_current_inventory_csv(
            inventory_xlsx=inventory_xlsx,
            inventory_csv=inventory_csv,
            include_none=include_none,
            lang=output_lang,
            verbose=verbose,
        )

    if exported:
        return exported

    checked = ", ".join(str(path) for path in checked_paths)
    raise FileNotFoundError(
        t("excel.error.no_inventory_excels_found", output_lang, checked=checked)
    )


def build_parser() -> argparse.ArgumentParser:
    root_language_parent = argparse.ArgumentParser(add_help=False)
    root_language_parent.add_argument(
        "--lang",
        choices=SUPPORTED_LANGUAGES,
        default=DEFAULT_LANG,
        help="CLI language: en (default) or uk.",
    )

    subcommand_language_parent = argparse.ArgumentParser(add_help=False)
    subcommand_language_parent.add_argument(
        "--lang",
        choices=SUPPORTED_LANGUAGES,
        default=argparse.SUPPRESS,
        help="CLI language: en (default) or uk.",
    )

    parser = argparse.ArgumentParser(
        description="Excel inventory by sheets and export back to current_inventory.<lang>.csv.",
        parents=[root_language_parent],
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    make_parser = subparsers.add_parser(
        "make-excel",
        help="Create an inventory Excel template from ingredients YAML files.",
        parents=[subcommand_language_parent],
    )
    make_parser.add_argument(
        "--ingredients-dir",
        default=str(DEFAULT_INGREDIENTS_DIR),
        help="Directory with ingredients_*.yaml files.",
    )
    make_parser.add_argument(
        "--inventory-xlsx",
        default=None,
        help="Target Excel path. If omitted, uses current_inventory.<lang>.xlsx.",
    )
    make_parser.add_argument(
        "--existing-csv",
        default=None,
        help="Existing current_inventory.<lang>.csv used to preserve entered amounts.",
    )

    export_parser = subparsers.add_parser(
        "export-csv",
        help="Export a filled inventory Excel file back to current_inventory.<lang>.csv.",
        parents=[subcommand_language_parent],
    )
    export_parser.add_argument(
        "--inventory-xlsx",
        default=None,
        help="Filled inventory Excel file. If omitted, exports all existing current_inventory.<lang>.xlsx files.",
    )
    export_parser.add_argument(
        "--inventory-csv",
        default=None,
        help="Target current_inventory.<lang>.csv path.",
    )
    export_parser.add_argument(
        "--only-existing",
        action="store_true",
        help="Write only rows where amount != none.",
    )

    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()
    args.lang = normalize_language(getattr(args, "lang", DEFAULT_LANG))

    service = InventoryExcelService(
        lang=args.lang,
        ingredients_dir=Path(getattr(args, "ingredients_dir", DEFAULT_INGREDIENTS_DIR)),
        inventory_xlsx=Path(args.inventory_xlsx) if getattr(args, "inventory_xlsx", None) else None,
        existing_csv=Path(args.existing_csv) if getattr(args, "existing_csv", None) else None,
        inventory_csv=Path(args.inventory_csv) if getattr(args, "inventory_csv", None) else None,
    )

    if args.command == "make-excel":
        service.make_inventory_excel()
    elif args.command == "export-csv":
        if args.inventory_xlsx or args.inventory_csv:
            service.export_excel_to_current_inventory_csv(include_none=not args.only_existing)
        else:
            export_all_default_inventory_excels_to_csv(
                include_none=not args.only_existing,
                output_lang=args.lang,
            )


if __name__ == "__main__":
    main()
